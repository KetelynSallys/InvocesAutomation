
import time
import pyautogui
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium import webdriver
import openpyxl
import os
from datetime import datetime
from openpyxl import Workbook
from selenium.common.exceptions import NoSuchElementException

ChromeDriverManager
Service

url_login = 'Insert your user'
url_site = 'Insert your URL site'
your_user = "Insert your user"
access_key = "Insert your password"

invoces_with_error = []

def load_dataframe(file_name="sheet.xlsx", sheet_name="Insert_your_sheet_page"):
    current_directory = os.path.dirname(os.path.abspath(__file__))
    file_path = os.path.join(current_directory, file_name)
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name]
        data = []
        headers = [cell.value for cell in sheet[1]]
        for row in sheet.iter_rows(min_row=2, values_only=True):data.append(dict(zip(headers, row)))
        num_items = len(data)
        print(f"Base with {num_items} CNPJ's")
        return data
    except FileNotFoundError:
        print(f"Error: File '{file_name}'not found.")
        return None
#print(f"ChromeDriver path: {chrome_options}")

def setup_driver():
    
    chrome_options = webdriver.ChromeOptions()
    chrome_options = Options()
    chrome_options.add_argument('--remote-debugging-port=9666')
    chrome_options.add_argument('--ignore-certificate-errors')
    driver = webdriver.Chrome( options=chrome_options)
    driver.set_window_size(1366, 768)
    return driver


def login(driver, url_login, your_user, access_key, url_site):
    try:
        driver.get(url_login)
        driver.find_element(By.XPATH, 'insert here your xpath referring to the user field').send_keys(your_user)
        driver.find_element(By.XPATH, 'insert here your xpath referring to the key field').send_keys(access_key)
        driver.find_element(By.XPATH, 'insert here your xpath referring to the button to enter').click()
        driver.get(url_site)
        iframe = driver.find_element(By.XPATH, 'insert your iframe or xpath of your search box element here')
        driver.switch_to.frame(iframe)
    except Exception as e:
        print(f"Error when you logging: {e}")


def search_cnpj(driver, cnpj):
    wait = WebDriverWait(driver, 30)
    try:
        element = wait.until(EC.element_to_be_clickable( (By.XPATH, 'insert here your xpath referring to the button top search')))
        element.clear()
        element.send_keys(cnpj)
        element = wait.until(EC.element_to_be_clickable((By.XPATH, 'insert here your xpath reffering result')))
        elemento.click()
        elemento = wait.until(EC.element_to_be_clickable((By.XPATH, 'insert here your xpath')))
        elemento.click()
        
        return True
    except Exception as e:
        print(f"Erro ao buscar CNPJ: {cnpj}")
        return False
    



def clicK_verify_download_invoice(driver, base_invoce_number,document):
    wait = WebDriverWait(driver, 30)
    
    table_rows= driver.find_elements(By.XPATH, f"insert here your xpath")
    quantity_of_items = len(table_rows)
    print(f"Number of records found: {quantity_of_items}")
    
    for i in range(2, quantity_of_items):
        
        xpath = f'insert here your xpath {i} generator'
        
        try:
           
            wait.until(EC.element_to_be_clickable((By.XPATH, xpath))).click()
            print(f"Button found {i}º registration")
            def get_number_note():
                invoce = WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, 'insert here your xpath')))
                numero_nota_web = invoce.text
                return numero_nota_web
            

            def formation_data(emission_data):
                        formation = ['%Y-%m-%d %H:%M:%S']
                        
                        for formation in formation:
                            try:
                                data = datetime.strptime(emission_data, formation)
                                return data.strftime("%d%m%Y")
            
                            except ValueError:
                            
                                raise ValueError(
                                    f"Unrecognized date/time format: {emission_data}")
            try:

                base_invoce_number = str(base_invoce_number).strip()
                document = str(document)
                invoce_number_web = get_number_note()
                formation_data()
                
                
                if int(invoce_number_web) == int(base_invoce_number):
                    print(str(f"Invoce {base_invoce_number} successfully found."))

                    time.sleep(3)
                    pyautogui.press('end')
                    time.sleep(3)
                    pyautogui.press('pagedown')
                    time.sleep(3)
                    pyautogui.click(704, 642)
                    time.sleep(15)
                    pyautogui.click(x=1238, y=183)
                    time.sleep(5)
                    pyautogui.press('enter')
                    time.sleep(5)
                    pyautogui.hotkey('ctrl', 'w')
                    time.sleep(5)

                    download_folder = r"insert here your path to repository"
                    filename_invoice = os.path.join(download_folder, "download.pdf")

                    new_filename_invoice = os.path.join(download_folder, f"DOC_{document}_NF_{base_invoce_number}.pdf")
                    os.rename(filename_invoice, new_filename_invoice)
                    print(f"File renamed to: {new_filename_invoice}")
                    return True
                else:
                    
                    print(str(f"Invoce number: {invoce_number_web}, does not match the desired grade: {base_invoce_number}"))
                    element = wait.until(EC.element_to_be_clickable((By.XPATH, 'insert here your xpath')))
                    element.click()

            except NoSuchElementException:
                print(f"Invoce unavailable: {base_invoce_number}")
                invoces_with_error.append([base_invoce_number])
                element = wait.until(EC.element_to_be_clickable((By.XPATH, 'insert here your xpath')))
                element.click()
                
        except Exception as e:
            print(f"Error finding the button {i}º registration: {e}")
           
            element = wait.until(EC.element_to_be_clickable((By.XPATH, 'insert here your xpath')))
            element.click()
        

def error_invoces():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "ERROR_INVOCES_UNAVAILABLE"
    for invoce in invoces_with_error:
        sheet.append([invoce])
        
    workbook.save("ERROR_INVOCES_UNAVAILABLE.xlsx")
    print("Error invoces saved successfully")

def main():
    while True:
        try:
            data = load_dataframe()
            if data is None:
                print("Error: Data is empty.")
                return
            driver = setup_driver()
            login(driver, url_login, your_user, access_key, url_site)
            for row in data:
                cnpj = row['CNPJ']
                base_invoice_number= row['INVOICE']
                document = row ['DOCUMENT']
                emission_data = row['EMISSION']
                search_cnpj(driver, cnpj)
                clicK_verify_download_invoice(
                driver, base_invoice_number,document)
        except Exception as e:
            print(f"Error executing function main: {e}")
           
if __name__ == '__main__':
    main()

    